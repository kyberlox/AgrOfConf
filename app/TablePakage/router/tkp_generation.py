# app/products/router/tkp_generation.py
import os
import re
from io import BytesIO
from typing import Optional, Union
from uuid import uuid4
from sqlalchemy import text

from fastapi import APIRouter, HTTPException, UploadFile, Depends, File, Form
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Mm 
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pathlib import Path
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from ..model.database import get_db
from ..model.tkp import TKP
from ..schema.tkp import TKPResponse
from datetime import datetime
from ..utils.router_utils import to_sql_name_lat

from app.StatisticsService.utils.deps import build_statistic_data
from app.UserService.utils.auth_utils import get_user_id_by_session_id
from app.StatisticsService.router.selection_router import get_selection_router

import requests
from openpyxl.drawing.image import Image as XLImage
from ..utils.kir_param_to_latin import KEY_MAPPING


router = APIRouter(prefix="/tkp_generation", tags=["TKP"])

UPLOAD_DIR = "./static/tkp_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Настройки
ALLOWED_EXTENSIONS = {".docx", ".xlsx"}



def validate_file(file: UploadFile) -> None:
    # Проверка расширения
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Invalid file extension. Allowed: .docx, .xlsx")

async def convert_data(user_dict: dict, db_info: dict) -> dict:
    data = datetime.strptime(db_info['date_search'], "%d.%m.%Y %H:%M:%S")
    user_dict['дата'] = data.strftime("%d.%m.%Y")
    user_dict['номер_запроса'] = user_dict['id']
    user_dict['адрес_исполнителя'] = db_info['user_work_city']
    user_dict['телефон_исполнителя'] = db_info['user_work_phone']
    user_dict['email_исполнителя'] = db_info['user_email']
    user_dict['фио_исполнителя'] = db_info['user_fio']
    user_dict['должность_исполнителя'] = db_info['user_work_position']
    return user_dict



# === TKP Generation Endpoints ===

@router.post("/create_tkp")
async def tkp_generation(
        file_id: int,
        product_id: int,
        user_dict: dict,
        db: AsyncSession = Depends(get_db),
        user_id: Optional[int] = Depends(get_user_id_by_session_id),
        statistic_router = Depends(get_selection_router),
):
    try:
        # Получаем файл из БД по id
        stmt = select(TKP).where(TKP.id == file_id)
        result = await db.execute(stmt)
        file_info = result.scalar_one_or_none()
        if not file_info:
            raise HTTPException(status_code=404, detail="Файл не найден")
        template_path = file_info.file
        contact_info = ["ФИО Заказчика", "Маркировка"]
        if not all(key in user_dict for key in contact_info):
            raise HTTPException(status_code=400, detail="Не все обязательные поля заполнены")

        filename = f"TKP_{to_sql_name_lat(user_dict['ФИО Заказчика'])}_{to_sql_name_lat(user_dict['Маркировка'])}"

        # Сохраняем статистику
        stat_info = await build_statistic_data(db, user_id, product_id)

        stat_info['parameters'] = user_dict

        document_number = await statistic_router.get_number_document(user_id)

        stat_info['document_number'] = document_number + 1

        is_dump = await statistic_router.save_selection(stat_info)

        user_dict['id'] = is_dump.data['elastic_response'].get("_id")

        user_dict = await convert_data(user_dict, stat_info)
        
        mark = user_dict.get("Маркировка")
        print(123)
        if mark:
            search_mark = mark[0:5]
            query = """
                SELECT file_path FROM product_drawing 
                WHERE product_id = :product_id 
                AND name = :name
            """ 
            params = {"product_id": product_id, "name": search_mark} 
            # Следить чтобы маркировка в БД и маркировка кодовая была одинаковой в плане кириллицы или латиницы
            stmt = await db.execute(text(query), params) 
            drawing_path = stmt.scalar_one_or_none()
        else:
            drawing_path = None
        print(123)
        if template_path.endswith(".docx"):
            print(123)
            doc = DocxTemplate(template_path)
            print(123)
            #Рендерим изображение
            if drawing_path:
                user_dict["Чертеж"] = InlineImage(doc, drawing_path, width=Mm(80)) 
            print(123)
            #Переводит на латиницу
            new_user_dict = {KEY_MAPPING[param]: value for param, value in user_dict.items()}
            print(new_user_dict)
            doc.render(new_user_dict)

            result_stream = BytesIO()
            doc.save(result_stream)
            result_stream.seek(0)

            return StreamingResponse(
                result_stream,
                media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}.docx"'
                }
            )

        elif template_path.endswith(".xlsx"):
            workbook = load_workbook(template_path, data_only=True)

            # for sheet in workbook.worksheets:
            #     for row in sheet.iter_rows():
            #         for cell in row:
            #             if isinstance(cell.value, str):
            #                 for key, value in user_dict.items():
            #                     pattern = re.compile(r'\{\{\s*' + re.escape(key) + r'\s*\}\}')
            #                     cell.value = pattern.sub(str(value), cell.value)
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            # Находим все плейсхолдеры в ячейке
                            pattern = re.compile(r'\{\{\s*([^}]+)\s*\}\}')
                            
                            def replace_match(match):
                                key = match.group(1).strip()
                                # Если ключ есть в словаре - возвращаем значение, иначе - пустую строку
                                return str(user_dict.get(key, ''))
                            
                            # Заменяем все плейсхолдеры
                            cell.value = pattern.sub(replace_match, cell.value)

             # Вставка изображения "Чертеж" на второй лист
            
            if len(workbook.worksheets) > 1 and drawing_path:
                try:
                    with open(drawing_path, 'rb') as file:
                        image_data = BytesIO(file.read())
                    
                    # Теперь файл закрыт, но данные сохранены в BytesIO
                    img = XLImage(image_data)
                    max_width = 400
                    max_height = 300
                    if img.width > max_width or img.height > max_height:
                        ratio = min(max_width / img.width, max_height / img.height)
                        img.width = int(img.width * ratio)
                        img.height = int(img.height * ratio)
                    # Якорь на ячейку A1 второго листа
                    img.anchor = 'A1'
                    second_sheet = workbook.worksheets[1]
                    second_sheet.add_image(img)
                except Exception as img_err:
                    # Если не удалось загрузить изображение — просто пропускаем
                    print(f"Не удалось вставить изображение: {img_err}")
            else:
                print('Не найден файл по заданной маркировке')

            result_stream = BytesIO()
            workbook.save(result_stream)
            result_stream.seek(0)

            return StreamingResponse(
                result_stream,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}.xlsx"'
                }
            )
        else:
            raise HTTPException(
                status_code=400,
                detail="Неподдерживаемый формат для файла"
            )
    except HTTPException:
        raise 
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при генерации ТКП: {str(e)}" 
        )

@router.post("/create_history_tkp", status_code=201, description="Создание ТКП из истории")
async def create_tkp_from_history(
    file_id: int,
    node_id: Union[int, str],
    db: AsyncSession = Depends(get_db),
    statistic_router = Depends(get_selection_router),
):
    from copy import deepcopy
    try:
        # Получаем историю из БД
        user_history = await statistic_router.get_selection_by_id(node_id)
        if not user_history:
            raise HTTPException(status_code=404, detail="История не найдена")
        
        # Получаем файл из БД по id
        stmt = select(TKP).where(TKP.id == file_id)
        result = await db.execute(stmt)
        file_info = result.scalar_one_or_none()
        if not file_info:
            raise HTTPException(status_code=404, detail="Файл не найден")
        template_path = file_info.file

        user_dict = deepcopy(user_history['parameters'])

        filename = f"TKP_{to_sql_name_lat(user_dict['Имя агента'])}_{to_sql_name_lat(user_dict['Маркировка'])}"
        user_dict['id'] = node_id
        user_dict = await convert_data(user_dict, user_history)
        
        if template_path.endswith(".docx"):
            doc = DocxTemplate(template_path)

            doc.render(user_dict)

            result_stream = BytesIO()
            doc.save(result_stream)
            result_stream.seek(0)

            return StreamingResponse(
                result_stream,
                media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}.docx"'
                }
            )

        elif template_path.endswith(".xlsx"):
            workbook = load_workbook(template_path, data_only=True)

            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            for key, value in user_dict.items():
                                pattern = re.compile(r'\{\{\s*' + re.escape(key) + r'\s*\}\}')
                                cell.value = pattern.sub(str(value), cell.value)

            result_stream = BytesIO()
            workbook.save(result_stream)
            result_stream.seek(0)

            return StreamingResponse(
                result_stream,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}.xlsx"'
                }
            )
        else:
            raise HTTPException(
                status_code=400,
                detail="Неподдерживаемый формат для файла"
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка при получении истории: {str(e)}")

@router.post("/add", response_model=TKPResponse, status_code=201, description="Добавление шаблона ТКП.")
async def add_tkp_file(
        product_id: int = Form(...),
        filename: str = Form(...),
        file: UploadFile = File(...),
        db: AsyncSession = Depends(get_db)
):
    try:
        validate_file(file)
        safe_filename = f"{uuid4().hex}{Path(file.filename).suffix.lower()}"
        file_path = os.path.join(UPLOAD_DIR, safe_filename)
        with open(file_path, "wb") as f:
            f.write(await file.read())

        file_url = f"/api/files/tkp_files/{safe_filename}"

        tkp_sample = TKP(
            name=filename,
            file=file_path,
            file_url=file_url,
            product_id=product_id
        )

        db.add(tkp_sample)
        await db.commit()
        await db.refresh(tkp_sample)

        return tkp_sample
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при добавлении ТКП: {str(e)}")


@router.get("/get_tkp_of_product/{product_id}", response_model=list[TKPResponse], description="Выведение всех ТКП продукта из БД.")
async def get_tkp_file(
        product_id: int,
        skip: int = 0,
        limit: int = 100,
        db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(TKP)
        .where(TKP.product_id == product_id)
        .offset(skip)
        .limit(limit)
    )

    samples = result.scalars().all()

    # if not samples:
    #     raise HTTPException(
    #         status_code=404,
    #         detail="TKP templates not found"
    #     )

    return samples


@router.delete("/delete_all_tkp_of_product/{product_id}", description="Удаление шаблона ТКП.")
async def delete_tkp_file(
        product_id: int,
        db: AsyncSession = Depends(get_db)
):
    try:
        result = await db.execute(select(TKP).where(TKP.product_id == product_id))
        samples = result.scalars().all()

        if not samples:
            return HTTPException(status_code=404, detail="TKP templates not found")

        for sample in samples:
            if sample.file is not None and sample.file != "" and os.path.exists(sample.file):
                os.remove(sample.file)

            await db.delete(sample)
        await db.commit()
        return {
            "detail": "TKP templates deleted successfully",
            "deleted_count": len(samples)
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка удаления шаблонов ТКП продукта с id = {product_id}: {str(e)}")

@router.delete("/delete/{tkp_id}", description="Удаление шаблона ТКП.")
async def delete_tkp_file(
        tkp_id: int,
        db: AsyncSession = Depends(get_db)
):
    try:
        result = await db.execute(select(TKP).where(TKP.id == tkp_id))
        sample = result.scalar_one_or_none()

        if sample is None:
            return HTTPException(status_code=404, detail="TKP template not found")

        if sample.file is not None and sample.file != "" and os.path.exists(sample.file):
            os.remove(sample.file)

        await db.delete(sample)
        await db.commit()
        return True
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка удаления шаблона ТКП: {str(e)}")